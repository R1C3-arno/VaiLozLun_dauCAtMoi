import math
import numpy as np
import pprint
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Alignment, PatternFill

## input ban đầu, tự thêm tùy tụi mày nhé
class ExcelLogger:
    def __init__(self, filename="paper_main_log.xlsx"):
        self.filename = filename
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Log"

        self.ws.append([
            "n", "j",
            "constraint_iter", "inner_iter",
            "Q", "P", "k1", "Av", "theta",
            "ts", "CR", "TC",
            "dQ", "dP", "dk1", "dAv", "dtheta",
            "LOG"
        ])

    def log(self,
            n, j,
            constraint_iter, inner_iter,
            Q, P, k1, Av, theta,
            ts, CR, TC,
            dQ="", dP="", dk1="", dAv="", dtheta="",
            log_msg=""):

        self.ws.append([
            n, j,
            constraint_iter, inner_iter,
            Q, P, k1, Av, theta,
            ts, CR, TC,
            dQ, dP, dk1, dAv, dtheta,
            log_msg
        ])

    def save(self):
        try:
            self.format_header()
            self.auto_adjust_column_width()
            self.colorize_columns()
            self.wb.save(self.filename)
            print(f"Log: Excel saved to: {self.filename}")
        except PermissionError:
            print("Log: Không lưu được Excel ")


    def auto_adjust_column_width(self):
        for col in self.ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            self.ws.column_dimensions[col_letter].width = max_length + 3

    def colorize_columns(self):
        colors = [
            "9CC6DB",  # xanh nhạt
            "FEEE91",  # vàng
            "DDBA7D",  # cam nhạt
            "F3E2D4",  # da
            "92487A",  # tím đậm
            "E49BA6",  # tím nhạt
            "415E72",  # xanh đậm
            "C5B0CD",  # khoai môn
        ]

        for col_idx in range(1, self.ws.max_column + 1):
            color = colors[(col_idx - 1) % len(colors)]
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            for row_idx in range(1, self.ws.max_row + 1):
                self.ws.cell(row=row_idx, column=col_idx).fill = fill

    def format_header(self):
        header_font = Font(bold=True, size=12)  #
        header_align = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(
            start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
        )

        for col in range(1, self.ws.max_column + 1):
            cell = self.ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = header_align
            cell.fill = header_fill


        self.ws.row_dimensions[1].height = 25

params = {
    "D": 600.0,
    "Pmin": 700.0,
    "Pmax": 1200.0,
    "Av0": 400.0,
    "A0": 150.0,
    "CT": 100.0,
    "sigma": 5.0,  ##nó là cái σ
    "rho": 1.0,  ## nó là cái ρ
    "hb": 19.234,
    "hv": 1.7,
    "pi": 80.0,  ## nó là π
    "xi1": 1.0 / 300.0,  ##ξ1 nó là cái này nè ξ1
    "xi2": 1.0 / 300.0,  ## nó là cái này nè ξ2
    "B1": 10000.0,
    "B2": 20.0,
    "theta0": 0.0002,  ## nó là cái này nè θ0
    "tT": 1.9,
}

a = [0.1, 0.15, 0.1]  ## cái set a1 a2 ae3
b = [0.05, 0.08, 0.04]  ## cái set b1 b2 b3
C = [10.0, 30.0, 70.0]  ## set C1 C2 C3
m = len(a)




m = len(a)

def compute_ts(a, b):
    m = len(a)
    ts_max = sum(a)
    ts_min = sum(b)
    ts_list = []

    for j in range(1, m + 1):  # j = 1..m (đúng như paper)
        sum_a = sum(a[i] for i in range(j, m))  # i từ j đến m-1 (trong Python)
        sum_b = sum(b[i] for i in range(0, j))  # i từ 0 đến j-1
        ts_j = sum_a - sum_b

        ## check miền b<ts<a
        if ts_min <= ts_j <= ts_max:
            ts_list.append(ts_j)
        else:
            print(f"Log: Error Bỏ ts_j={ts_j} (ngoài [{ts_min}, {ts_max}])")

    return ts_list


def compute_CR(ts_list, a, b, C):
    m = len(a)
    ts_max = sum(a)

    CR = []

    #sort Cr theo thứ tự tăng dần
    zip_all = sorted(zip(a, b, C), key=lambda x: x[2])
    a, b, C = map(list, zip(*zip_all))

    for j in range(1, len(ts_list) + 1):  # j = 1..m
        ts_prev = ts_list[j - 2] if j > 1 else ts_max  # t_{s,j-1}
        ts_curr = ts_list[j - 1]  # t_{s,j}

        term1 = C[j - 1] * max(ts_prev - ts_curr, 0)

        # sum i = 1 .. min(j+1, m)  (chú ý: Python index = i-1)
        last_i = min(j + 1, m)
        term2 = 0.0
        for i in range(1, last_i+1):  # i from 1 to j+1
            idx = i - 1  # python index
            term2 += C[idx] * (a[idx] - b[idx])

        CR.append(term1 + term2)

    return CR


def compute_TCv(Q, P, n, Av, theta, params):
    D = params["D"]
    hv = params["hv"]
    B1 = params["B1"]
    B2 = params["B2"]
    theta0 = params["theta0"]
    rho = params["rho"]
    Av0 = params["Av0"]
    xi1 = params["xi1"]
    xi2 = params["xi2"]


    CP = xi1 * P + xi2 / P

    T1 = Av * D / (n * Q)

    T2 = hv * Q / 2 * (
            n * (1 - D / P) - 1 + 2 * D / P
    )

    T3 = D * CP

    T4 = B1 * math.log(theta / theta0)

    T5 = B2 * math.log(Av0 / Av)

    T6 = rho * D * theta * n * Q / 2

    return T1 + T2 + T3 + T4 + T5 + T6


def compute_TCb(Q, P, n, k1, ts, tT, CR_ts, params):
    D = params["D"]
    A0 = params["A0"]
    CT = params["CT"]
    hb = params["hb"]
    sigma = params["sigma"]
    pi = params["pi"]

    ts_safe = max(ts + Q / P, 1e-10)

    # --------- Term 1 ----------
    T1 = (A0 + n * CT) * D / (n * Q)

    # --------- Term 2 ----------
    T2 = hb * (Q / 2 + k1 * sigma * math.sqrt(ts_safe))

    # --------- E(x1 - R1)+ ----------
    Ex1 = (sigma / 2) * math.sqrt(ts_safe) * \
          (math.sqrt(1 + k1 ** 2) - k1)

    # --------- E(x2 - R2)+ ----------
    Ex2 = (sigma / 2) * math.sqrt(tT) * (
            math.sqrt(1 + k1 ** 2 * ts_safe / tT)
            - k1 * math.sqrt(ts_safe / tT)
    )

    T3 = D * pi / (n * Q) * Ex1
    T4 = D * pi * (n - 1) / (n * Q) * Ex2

    # --------- Crashing cost ----------
    T5 = D * n * CR_ts / (n * Q)

    return T1 + T2 + T3 + T4 + T5


def compute_TC_total(Q, P, n, k1, Av, theta, params, ts, tT, CR_ts):
    TCb = compute_TCb(Q, P, n, k1, ts, tT, CR_ts, params)
    TCv = compute_TCv(Q, P, n, Av, theta, params)
    return TCb + TCv


def daoham_Tcb_theoQ(Q, P, n, k1, ts, tT, CR_ts, params):

    D = params["D"]
    A0 = params["A0"]
    CT = params["CT"]
    hb = params["hb"]
    sigma = params["sigma"]
    pi = params["pi"]

    ts_safe = max(ts + Q / P, 1e-10)
    # đạo hàm  D(A0+nCt)/nQ
    term1 = -((A0 + n * CT) * D) / (n * Q ** 2)
    # đạo hàm  hb/2ksigma/2Psqrt(ts+Q/p
    term2 = (hb / 2) + (hb * k1 * sigma) / (2 * P * math.sqrt(ts_safe))
    # đạo hàm
    term3_1 = (D * pi * sigma * (math.sqrt(1 + k1 ** 2) - k1)) / (4 * n * P * Q * math.sqrt(ts_safe))
    term3_2 = ((D * pi) / (n * Q ** 2)) * (sigma / 2) * (math.sqrt(ts_safe) * (math.sqrt(1 + k1 ** 2) - k1))
    term3 = term3_1 - term3_2

    # Ex2 gốc
    Ex2 = (sigma / 2) * math.sqrt(tT) * (
            math.sqrt(1 + k1 ** 2 * ts_safe / tT)
            - k1 * math.sqrt(ts_safe / tT)
    )

    # đạo hàm Ex2 theo Q
    u = ts_safe
    dEx2_dQ = (sigma / (2 * P)) * (
            (k1 ** 2) / (2 * math.sqrt(tT) * math.sqrt(1 + k1 ** 2 * u / tT))
            - k1 / (2 * math.sqrt(u))
    )

    term4 = (
            - (D * pi * (n - 1)) / (n * Q ** 2) * Ex2
            + (D * pi * (n - 1)) / (n * Q) * dEx2_dQ
    )

    term5 = -D * CR_ts / (Q ** 2)

    daoham_tcb = term1 + term2 + term3 + term4 + term5
    return daoham_tcb


def daoham_Tcv_theoQ(Q, P, n, Av, theta, params):
    D = params["D"]
    hv = params["hv"]
    rho = params["rho"]

    dT1 = - Av * D / (n * Q ** 2)

    const = n * (1 - D / P) - 1 + 2 * D / P
    dT2 = hv / 2 * const

    dT6 = rho * D * theta * n / 2

    return dT1 + dT2 + dT6


def b1(D, n, A0, CT, pi, sigma, ts, Q, P, tT, k1, CR, Av):
    ts_safe = max(ts + Q / P, 1e-8)

    term1 = (sigma / 2) * math.sqrt(ts_safe) * (math.sqrt(1 + k1 ** 2) - k1)

    term2 = ((n - 1) * sigma / 2) * math.sqrt(tT + k1 ** 2 * ts_safe)

    term3 = n * CR / pi

    return (D / n) * (A0 + n * CT + pi * (term1 + term2 + term3) + Av)


def b2(D, pi, sigma, ts, Q, P, tT, k1, n):

    ts_safe = max(ts + Q / P, 1e-8)

    tT_safe = tT + k1 ** 2 * ts_safe


    term1 = (math.sqrt(1 + k1 ** 2) - k1) / math.sqrt(ts_safe)

    term2_1 = -1 / math.sqrt(ts_safe)
    term2 = (n - 1) * k1 * (term2_1 + k1 / math.sqrt(tT_safe))

    return (D * pi * sigma) / (4 * n * P) * (term1 + term2)


def b3(hb, sigma, k1, ts, Q, P, hv, n, D, rho, theta):
    ts_safe = ts + Q / P


    term1 = hb * (0.5 + (k1 * sigma) / (2 * P * math.sqrt(ts_safe)))

    term2 = (hv / 2) * (n * (1 - D / P) - 1 + 2 * D / P)

    term3 = (rho * n * D * theta) / 2

    return term1 + term2 + term3


## hàm này tính Q* nè mấy thằng lồn

def giai_nghiem_daohamTc_theoQ(Q0, P, n, Av, theta, ts, tT, k1, CR, params):
    D = params["D"]
    A0 = params["A0"]
    CT = params["CT"]
    hb = params["hb"]
    sigma = params["sigma"]
    pi = params["pi"]
    hv = params["hv"]
    rho = params["rho"]

    Q = Q0
    for _ in range(100):
        beta1 = b1(D, n, A0, CT, pi, sigma, ts, Q, P, tT, k1, CR, Av)
        beta2 = b2(D, pi, sigma, ts, Q, P, tT, k1, n)
        beta3 = b3(hb, sigma, k1, ts, Q, P, hv, n, D, rho, theta)

        Q_new = beta1 / (beta2 + beta3 * Q)
        if not np.isfinite(Q_new):
            Q_new = Q
        Q_new = max(Q_new, 1e-3)

        ## điều kiện hội tụ ( kèm Q_new vào ko nó nhảy sập lồn)
        if abs(Q_new - Q) / max(Q, 1e-8) < 1e-6:
            return Q_new
        Q = Q_new

    ## không hội tụ trả về Q=Q0
    return Q

def b4(ts, Q, P, n, hv, tT, k1, params):
    D = params["D"]
    hb = params["hb"]
    sigma = params["sigma"]
    pi = params["pi"]
    xi2 = params["xi2"]

    ts_safe = ts + Q / P
    ts_sqrt = math.sqrt(ts_safe)

    tT_safe = tT + k1 ** 2 * ts_safe
    tT_sqrt = math.sqrt(tT_safe)

    term1 = hb * k1 * sigma * Q / (2 * ts_sqrt)

    trong_ngoac_1 = (math.sqrt(1 + k1 ** 2) - k1) / ts_sqrt
    trong_ngoac_2 = (n - 1) * (k1 /tT_sqrt - 1 / ts_sqrt)

    term2 = (D * pi * sigma / (4 * n)) * (trong_ngoac_1 + trong_ngoac_2)

    term3 = - hv * Q * D * (n - 2) / 2
    term4 = D * xi2

    return term1 + term2 + term3 + term4


def daoham_Tc_theoP(P, b4_value, params):
    D = params["D"]
    xi1 = params["xi1"]

    return - b4_value / (P ** 2) + D * xi1


## tính P* nè mấy con gà
def giai_nghiem_daohamTc_theoP(b4, P_old, params, alpha=0.3, eps=1e-12):


    D = params["D"]
    xi1 = params["xi1"]
    Pmin = params["Pmin"]
    Pmax = params["Pmax"]

    logs = []

    val = b4 / (D * xi1)

    if val <= 0 or not math.isfinite(val):
        #  KHÔNG cho rơi về Pmin nữa
        P_fp = P_old
        logs.append(f"SKIP_P_B4_INVALID ({val:.3e})")
    else:
        P_fp = math.sqrt(val)


    # RELAXATION (nới mềm)

    P_relaxed = (1 - alpha) * P_old + alpha * P_fp


    return P_relaxed, logs


## đạo hàm TC theo k1
def daoham_Tc_theoK1(k1, Q, P, n, ts, tT, params):
    D = params["D"]
    hb = params["hb"]
    sigma = params["sigma"]
    pi = params["pi"]

    ts_safe = ts + Q / P
    tT_safe = tT + k1 ** 2 * ts_safe

    term1_1 = k1 / math.sqrt(1 + k1 ** 2) - 1
    term1_2 = (n - 1) * (k1 * math.sqrt(ts_safe / tT_safe) - 1)

    term1 = hb + (pi * D / (2 * n * Q)) * (term1_1 + term1_2)

    term2 = sigma * math.sqrt(ts_safe)

    return term2 * term1


## hàm tính t* (tử chia mẫu)

def capnhat_k1_theo_congthuc(k1, Q, P, n, ts, tT, params):
    D = params["D"]
    hb = params["hb"]
    pi = params["pi"]

    u = ts + Q / P

    v = tT + k1 ** 2 * u

    tu = n * (D * pi - 2 * Q * hb)

    mau = D * pi * (
            1 / math.sqrt(1 + k1 ** 2)
            + (n - 1) * math.sqrt(u / v)
    )

    k1_new = tu / mau

    return k1_new


##hàm này tìm k1*

def giai_nghiem_daohamTc_theoK1(Q, P, n, ts, tT, k1_old, params):


    k1 = capnhat_k1_theo_congthuc(k1_old, Q, P, n, ts, tT, params)

    return k1


##hàm này tính Av *
def giai_nghiem_daohamTc_theoAv(Q, n, params):
    D = params["D"]
    B2 = params["B2"]
    return n * Q * B2 / D


## hàm này giải nghiệm theta*
def giai_nghiem_daohamTc_theoTheta(Q, n, params):
    B1 = params["B1"]
    rho = params["rho"]
    D = params["D"]
    return 2 * B1 / (rho * n * D * Q)

def ep_bien(Q, P, k1, Av, theta, params, eps=1e-8,alpha=0.39173):
    logs = []

    # ---- Q ----
    if Q < eps:
        logs.append(f"CLIP_Q ({Q:.3e} -> {eps:.3e})")
        Q = eps

    # ---- P ----
    Pmin, Pmax = params["Pmin"], params["Pmax"]
    if P < Pmin:
        P_old = P
        P = Pmin + alpha * (Pmin - P_old)
        logs.append(f"SOFT_CLIP_P_LOW ({P_old:.3e} -> {P:.3e})")

    elif P > Pmax:
        P_old = P
        P = Pmax + alpha * (P_old - Pmax)
        logs.append(f"SOFT_CLIP_P_HIGH ({P_old:.3e} -> {P:.3e})")

    # ---- k1 ----
    if k1 < eps:
        logs.append(f"CLIP_k1 ({k1:.3e} -> {eps:.3e})")
        k1 = eps


    return Q, P, k1, Av, theta, logs




# === MAIN: ===

def paper_main(params: Dict, a: List[float], b: List[float], C: List[float],
               max_n_search: int = 10,
               max_inner_iter: int = 10,
               max_constraint_iter: int = 10,
               tol: float = 10):
    logger = ExcelLogger("paper_debug.xlsx")

    # =========================
    # STEP 1: compute ts_list, CR_list
    # =========================
    ts_list = compute_ts(a, b)
    CR_list = compute_CR(ts_list, a, b, C)

    print("STEP 1: ts_list =", ts_list)
    print("STEP 1: CR_list =", CR_list)

    all_solutions = []
    TC_prev_n = float("inf")


    # STEP 5: Outer loop on n

    n = 1
    while n <= max_n_search:
        print("\n" + "=" * 80)
        print(f"STEP 5: Trying n = {n}")
        print("=" * 80)

        TC_best_for_n = float("inf")
        best_solution_for_n = None


        # STEP 2: Loop over ts_j

        for j_idx, ts in enumerate(ts_list, start=1):
            CR_ts = CR_list[j_idx - 1]

            print("\n" + "-" * 60)
            print(f"STEP 2: n={n}, j={j_idx} | ts={ts:.6f}, CR={CR_ts:.6f}")
            print("-" * 60)


            # INITIALIZATION

            Q = 200.0
            P = 900.0
            k1 = 4.0
            Av = params["Av0"]
            theta = params["theta0"]

            lock_Av = False
            lock_theta = False
            feasible = False


            # CONSTRAINT LOOP (paper STEP 3)

            for constraint_iter in range(max_constraint_iter):

                converged = False


                # INNER FIXED-POINT LOOP (STEP 2b–2e)
                # Giải nghiệm độc lập

                for iter_idx in range(max_inner_iter):
                    Q_old, P_old, k1_old, Av_old, theta_old = Q, P, k1, Av, theta
                    clip_logs = []


                    # --- 1. Update P (độc lập) ---
                    b4_val = b4(ts, Q_old, P_old, n, params["hv"], params["tT"], k1_old, params)
                    P_new, P_logs = giai_nghiem_daohamTc_theoP(b4_val, P_old, params, alpha=0.25)
                    clip_logs.extend(P_logs)

                    # --- 2. Update Q (độc lập) ---
                    Q_new = giai_nghiem_daohamTc_theoQ(
                        Q_old, P_old, n, Av_old, theta_old,
                        ts, params["tT"], k1_old, CR_ts, params
                    )

                    # --- 3. Update k1 (độc lập) ---
                    k1_new = giai_nghiem_daohamTc_theoK1(
                        Q_old, P_old, n, ts, params["tT"], k1_old, params
                    )

                    # --- 4. Update Av, theta (độc lập) ---
                    if lock_Av:
                        Av_new = Av
                    else:
                        Av_new = giai_nghiem_daohamTc_theoAv(Q_old, n, params)

                    if lock_theta:
                        theta_new = theta
                    else:
                        theta_new = giai_nghiem_daohamTc_theoTheta(Q_old, n, params)

                    # UPDATE + Ép biên
                    Q, P, k1, Av, theta,clip_logs = ep_bien(Q_new, P_new, k1_new, Av_new, theta_new,params)

                    for msg in clip_logs:
                        logger.log(
                            n=n, j=j_idx,
                            constraint_iter=constraint_iter + 1,
                            inner_iter=iter_idx + 1,
                            Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                            ts=ts, CR=CR_ts, TC="",
                            log_msg=msg
                        )

                    TC_val_iter = compute_TC_total(Q, P, n, k1, Av, theta, params, ts, params["tT"], CR_ts)


                    # CHECK HỘI TỤ - Giải fixed point

                    dQ = abs(Q - Q_old)
                    dP = abs(P - P_old)
                    dk1 = abs(k1 - k1_old)
                    dAv = abs(Av - Av_old)
                    dtheta = abs(theta - theta_old)



                    logger.log(
                        n=n, j=j_idx,
                        constraint_iter=constraint_iter + 1,
                        inner_iter=iter_idx + 1,
                        Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                        ts=ts, CR=CR_ts, TC=TC_val_iter,
                        dQ=dQ, dP=dP, dk1=dk1, dAv=dAv, dtheta=dtheta,
                        log_msg="INNER_STEP"
                    )

                    if max(dQ, dP, dk1, dAv, dtheta) < tol:
                        converged = True
                        print(f"   Log: Inner converged at iter {iter_idx + 1}")
                        logger.log(
                            n=n, j=j_idx,
                            constraint_iter=constraint_iter + 1,
                            inner_iter=iter_idx + 1,
                            Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                            ts=ts, CR=CR_ts, TC=TC_val_iter,
                            dQ=dQ, dP=dP, dk1=dk1, dAv=dAv, dtheta=dtheta,
                            log_msg="INNER_STEP"
                        )

                        break

                if not converged:
                    print("  Log:  Error: Inner loop NOT converged (paper vẫn cho đi tiếp)")

                print(f"  Log:  After inner: Q={Q:.6f}, P={P:.6f}, k1={k1:.6f}, Av={Av:.6f}, theta={theta:.6e}")

                # =========================
                # CHECK RÀNG BUỘC PAPER
                # =========================
                Av_violate = Av > params["Av0"]
                theta_violate = theta > params["theta0"]

                if (not Av_violate) and (not theta_violate):
                    feasible = True
                    # log constraint ok (use last computed TC_val_iter)
                    logger.log(
                        n=n, j=j_idx,
                        constraint_iter=constraint_iter + 1,
                        inner_iter="",
                        Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                        ts=ts, CR=CR_ts, TC=TC_val_iter,
                        log_msg="CONSTRAINT_OK"
                    )
                    print("    Log: OK Constraints satisfied")
                    break

                if Av_violate:
                    Av = params["Av0"]
                    lock_Av = True

                    TC_post_reset = compute_TC_total(Q, P, n, k1, Av, theta, params, ts, params["tT"], CR_ts)
                    logger.log(
                        n=n, j=j_idx,
                        constraint_iter=constraint_iter + 1,
                        inner_iter="",
                        Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                        ts=ts, CR=CR_ts, TC=TC_post_reset,
                        log_msg="RESET_Av"
                    )
                    print("   Log:Error:  Av violated → set Av = Av0")

                if theta_violate:
                    theta = params["theta0"]
                    lock_theta = True  #

                    TC_post_reset = compute_TC_total(Q, P, n, k1, Av, theta, params, ts, params["tT"], CR_ts)
                    logger.log(
                        n=n, j=j_idx,
                        constraint_iter=constraint_iter + 1,
                        inner_iter="",
                        Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                        ts=ts, CR=CR_ts, TC=TC_post_reset,
                        log_msg="RESET_theta"
                    )
                    print("    Log:Error theta violated → set theta = theta0")

            # =========================
            # Nếu không feasible → bỏ j
            # =========================
            if not feasible:
                print("  Log: Error: Infeasible → discard j")
                logger.log(
                    n=n, j=j_idx,
                    constraint_iter="", inner_iter="",
                    Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                    ts=ts, CR=CR_ts, TC="",
                    log_msg="DISCARD_j"
                )

                continue


            # STEP 4: Compute TC

            TC_val = compute_TC_total(Q, P, n, k1, Av, theta, params, ts, params["tT"], CR_ts)

            logger.log(
                n=n,
                j=j_idx,
                constraint_iter=constraint_iter + 1,
                inner_iter=iter_idx + 1,
                Q=Q, P=P, k1=k1, Av=Av, theta=theta,
                ts=ts, CR=CR_ts, TC=TC_val,
                dQ=dQ, dP=dP, dk1=dk1, dAv=dAv, dtheta=dtheta,
                log_msg="INNER_STEP"
            )

            print(f"  -> TC(n={n}, j={j_idx}) = {TC_val:.6f}")

            if TC_val < TC_best_for_n:
                TC_best_for_n = TC_val
                best_solution_for_n = {
                    "n": n,
                    "j": j_idx,
                    "Q": Q,
                    "P": P,
                    "k1": k1,
                    "Av": Av,
                    "theta": theta,
                    "ts": ts,
                    "CR": CR_ts,
                    "TC": TC_val
                }


        # STEP 5: So sánh TC theo n

        print("\n" + "=" * 60)
        print(f"BEST TC for n={n} = {TC_best_for_n}")
        print(f"PREVIOUS TC (n-1) = {TC_prev_n}")

        if TC_best_for_n < TC_prev_n:
            TC_prev_n = TC_best_for_n
            all_solutions.append(best_solution_for_n)
            n += 1
            logger.log(
                n=n, j="",
                constraint_iter="", inner_iter="",
                Q="", P="", k1="", Av="", theta="",
                ts="", CR="", TC="",
                log_msg="STOP_BY_TC"
            )

        else:
            # Log stopping reason with current n (do NOT increment n)
            logger.log(
                n=n, j="",
                constraint_iter="", inner_iter="",
                Q="", P="", k1="", Av="", theta="",
                ts="", CR="", TC="",
                log_msg="STOP_BY_TC"
            )
            print("Log: Error: TC không giảm nữa → DỪNG đúng paper")
            break


    # STEP 6: Lấy nghiệm tốt nhất

    print("\n" + "=" * 80)
    print("STEP 6: FINAL RESULT")
    logger.save()

    if all_solutions:
        best_overall = min(all_solutions, key=lambda x: x["TC"])
        pprint.pprint(best_overall, indent=2)
        return best_overall, all_solutions
    else:
        print("Log:Error: Không có nghiệm feasible")
        return None, all_solutions

if __name__ == "__main__":
    # Run the main routine
    best_sol, all_sols = paper_main(params, a, b, C,
                                    max_n_search=50,
                                    max_inner_iter=200,
                                    max_constraint_iter=200,
                                    tol=1e-9)

    if best_sol:
        print("\nLOG: OK FINAL OPTIMUM")
        print(f"n* = {best_sol['n']}")
        print(f"j* = {best_sol['j']}")
        print(f"Q* = {best_sol['Q']:.6f}")
        print(f"P* = {best_sol['P']:.6f}")
        print(f"k1* = {best_sol['k1']:.6f}")
        print(f"Av* = {best_sol['Av']:.6f}")
        print(f"theta* = {best_sol['theta']:.9e}")
        print(f"TC* = {best_sol['TC']:.6f}")
    else:
        print("No feasible solution produced by algorithm.")